"""
Router de Exportações — /api/v1/exports
Gera arquivos CSV, XLSX e JSON via Celery (async).

Fluxo:
  POST /exports/candidatos   → Celery task → retorna job_id
  GET  /exports/{job_id}     → status e URL de download quando pronto
  GET  /exports/{job_id}/download → stream do arquivo gerado
"""
from __future__ import annotations

import io
import logging
import os
import uuid
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from db.session import get_db
from shared.schemas import ExportJobOut, SourceMeta

router = APIRouter()
logger = logging.getLogger(__name__)

# Em produção, usar MinIO/S3. Em dev, armazena in-memory.
_EXPORT_CACHE: Dict[str, Dict[str, Any]] = {}

_SOURCE_TSE = SourceMeta(
    fonte="TSE — Portal de Dados Abertos",
    url="https://dadosabertos.tse.jus.br/",
    cobertura_temporal="2000–2024",
    cobertura_geografica="Nacional",
)


# ─────────────────────────────────────────────
# Helpers de geração de arquivo
# ─────────────────────────────────────────────

def _rows_to_csv(rows: List[Dict[str, Any]]) -> bytes:
    import csv, io
    if not rows:
        return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM para Excel


def _rows_to_xlsx(rows: List[Dict[str, Any]], sheet_name: str = "Dados") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl não instalado — retornando CSV")
        return _rows_to_csv(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(rows[0].keys())

    # Header formatado
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))
        # Zebragem
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="EFF6FF")

    # Largura automática
    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), max((len(str(row.get(header, ""))) for row in rows[:100]), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Linha de rodapé com fonte
    footer_row = len(rows) + 3
    ws.cell(row=footer_row, column=1, value=f"Fonte: TSE — dadosabertos.tse.jus.br | Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="6B7280")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# Geração assíncrona de exportações
# ─────────────────────────────────────────────

async def _generate_candidatos_export(
    job_id: str,
    election_id: Optional[str],
    uf: Optional[str],
    cargo_codigo: Optional[str],
    partido: Optional[str],
    situacao: Optional[str],
    formato: str,
):
    """Gera exportação de candidatos e armazena no cache."""
    from db.models.political import Candidacy, Party, Office, Election
    from db.models.territorial import Territory
    from db.session import AsyncSessionLocal

    _EXPORT_CACHE[job_id]["status"] = "running"

    try:
        async with AsyncSessionLocal() as session:
            stmt = (
                select(Candidacy)
                .options(
                    selectinload(Candidacy.person),
                    selectinload(Candidacy.party),
                    selectinload(Candidacy.office),
                    selectinload(Candidacy.territory),
                    selectinload(Candidacy.election),
                )
                .limit(50_000)
            )

            if election_id:
                stmt = stmt.where(Candidacy.election_id == uuid.UUID(election_id))
            if situacao:
                stmt = stmt.where(Candidacy.situacao.ilike(f"%{situacao}%"))
            if cargo_codigo:
                stmt = stmt.join(Office, Office.id == Candidacy.office_id).where(Office.codigo == cargo_codigo)
            if partido:
                stmt = stmt.join(Party, Party.id == Candidacy.party_id).where(Party.sigla.ilike(f"%{partido}%"))
            if uf:
                stmt = stmt.join(Territory, Territory.id == Candidacy.territory_id).where(Territory.uf == uf.upper())

            rows_db = (await session.execute(stmt)).scalars().all()

        rows = [
            {
                "nome": c.person.nome if c.person else "",
                "nome_urna": c.person.nome_urna if c.person else "",
                "partido": c.party.sigla if c.party else "",
                "cargo": c.office.descricao if c.office else "",
                "municipio": c.territory.nome if c.territory else "",
                "uf": c.territory.uf if c.territory else "",
                "eleicao_ano": c.election.ano if c.election else "",
                "turno": c.election.turno if c.election else "",
                "situacao": c.situacao,
                "numero_urna": c.numero_urna,
                "votos_totais": c.votos_totais,
                "bens_declarados_total": float(c.bens_declarados_total) if c.bens_declarados_total else None,
                "genero": c.person.genero if c.person else "",
                "raca_cor": c.person.raca_cor if c.person else "",
                "escolaridade": c.person.escolaridade if c.person else "",
                "ocupacao": c.person.ocupacao if c.person else "",
                "nascimento": c.person.nascimento.isoformat() if c.person and c.person.nascimento else "",
                "fonte_url": c.fonte_url,
            }
            for c in rows_db
        ]

        if formato == "xlsx":
            content = _rows_to_xlsx(rows, "Candidatos")
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext = "xlsx"
        elif formato == "json":
            import json
            content = json.dumps({"data": rows, "total": len(rows), "fonte": "TSE"}, ensure_ascii=False, default=str).encode("utf-8")
            media_type = "application/json"
            ext = "json"
        else:
            content = _rows_to_csv(rows)
            media_type = "text/csv"
            ext = "csv"

        expires = datetime.now(timezone.utc) + timedelta(hours=24)
        _EXPORT_CACHE[job_id].update({
            "status": "done",
            "content": content,
            "media_type": media_type,
            "filename": f"candidatos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}",
            "total_rows": len(rows),
            "expires_at": expires.isoformat(),
        })

    except Exception as e:
        logger.error(f"Export error [{job_id}]: {e}", exc_info=True)
        _EXPORT_CACHE[job_id]["status"] = "failed"
        _EXPORT_CACHE[job_id]["error"] = str(e)


# ─────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────

@router.post("/exports/candidatos", response_model=ExportJobOut)
async def export_candidatos(
    background_tasks: BackgroundTasks,
    election_id: Optional[str] = Query(None),
    uf: Optional[str] = Query(None),
    cargo_codigo: Optional[str] = Query(None),
    partido: Optional[str] = Query(None),
    situacao: Optional[str] = Query(None),
    formato: str = Query("csv", regex="^(csv|xlsx|json)$"),
) -> ExportJobOut:
    """
    Inicia exportação de candidatos com filtros.
    Retorna imediatamente com job_id para polling.
    """
    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)

    _EXPORT_CACHE[job_id] = {
        "status": "pending",
        "formato": formato,
        "created_at": now.isoformat(),
    }

    background_tasks.add_task(
        _generate_candidatos_export,
        job_id=job_id,
        election_id=election_id,
        uf=uf,
        cargo_codigo=cargo_codigo,
        partido=partido,
        situacao=situacao,
        formato=formato,
    )

    return ExportJobOut(
        job_id=job_id,
        status="pending",
        formato=formato,
        created_at=now,
    )


@router.get("/exports/{job_id}", response_model=ExportJobOut)
async def get_export_status(job_id: str) -> ExportJobOut:
    """Status de uma exportação. Quando done, inclui url_download."""
    job = _EXPORT_CACHE.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado ou expirado")

    created = datetime.fromisoformat(job["created_at"]) if "created_at" in job else None
    expires = datetime.fromisoformat(job["expires_at"]) if "expires_at" in job else None

    url = f"/api/v1/exports/{job_id}/download" if job["status"] == "done" else None

    return ExportJobOut(
        job_id=job_id,
        status=job["status"],
        formato=job.get("formato", "csv"),
        url_download=url,
        expires_at=expires,
        created_at=created,
    )


@router.get("/exports/{job_id}/download")
async def download_export(job_id: str) -> StreamingResponse:
    """Faz download do arquivo de exportação gerado."""
    job = _EXPORT_CACHE.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado ou expirado")
    if job["status"] != "done":
        raise HTTPException(status_code=202, detail=f"Job status: {job['status']}")

    content = job.get("content", b"")
    filename = job.get("filename", "export.csv")
    media_type = job.get("media_type", "text/csv")

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Total-Rows": str(job.get("total_rows", 0)),
            "X-Source": "TSE — dadosabertos.tse.jus.br",
        },
    )


@router.delete("/exports/{job_id}")
async def cancel_export(job_id: str):
    """Remove um job do cache (cleanup antecipado)."""
    if job_id in _EXPORT_CACHE:
        del _EXPORT_CACHE[job_id]
    return {"detail": "removed"}
